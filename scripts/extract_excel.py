#!/usr/bin/env python3
"""Extract all sheets from IT Asset New (3).xlsx into JSON for seeding."""
import json
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/home/z/my-project/upload/IT Asset New (3).xlsx")
OUT = Path("/home/z/my-project/scripts/excel_data.json")

wb = load_workbook(SRC, data_only=True, read_only=True)

TYPE_MAP = {
    "Desktop": "Desktop",
    "Laptop": "Laptop",
    "MobileTablet": "Mobile",
    "PrinterScanner": "Printer",
    "PDT": "PDT",
    "Other": "Other",
    "POS": "POS",
    "Bill printer": "Bill Printer",
    "Weighing Scale": "Weighing Scale",
    "Biometric": "Biometric",
    "NVR": "NVR",
    "Firewall": "Firewall",
    "Router": "Router",
    "Switch": "Switch",
}

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ("nan", "none", "n/a", "na", "null", "-"):
            return None
        return s
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, float):
        return v
    return v

def extract_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = 0
    for i, r in enumerate(rows[:5]):
        non_empty = sum(1 for c in r if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break
    headers = [clean(h) for h in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        non_empty = sum(1 for c in r if c is not None and str(c).strip())
        if non_empty == 0:
            continue
        row_obj = {}
        for h, v in zip(headers, r):
            if h:
                row_obj[h] = clean(v)
        if not row_obj:
            continue
        # permissive: at least one identifier
        id_cols = ["Make", "Model", "Model Number", "S/N", "Service Tag(S/N)",
                   "Serial Number", "Computer Name", "User", "Store Name"]
        if not any(row_obj.get(k) for k in id_cols):
            continue
        out.append(row_obj)
    return out

result = {}
for sheet_name in wb.sheetnames:
    if sheet_name not in TYPE_MAP:
        continue
    ws = wb[sheet_name]
    data = extract_sheet(ws)
    result[sheet_name] = {
        "assetType": TYPE_MAP[sheet_name],
        "headers": list(data[0].keys()) if data else [],
        "rows": data,
        "count": len(data),
    }

OUT.write_text(json.dumps(result, indent=2, default=str))
print(f"Wrote {sum(s['count'] for s in result.values())} rows across {len(result)} sheets to {OUT}")
for k, v in result.items():
    print(f"  {k}: {v['count']} rows, headers: {v['headers']}")
